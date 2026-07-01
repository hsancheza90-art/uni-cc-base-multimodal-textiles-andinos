from __future__ import annotations

import argparse
import csv
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = "data/metadata/met_revision_manual_v2.xlsx"
HOJA_REVISION = "revision_met_v2"

SALIDAS = {
    "principal": "data/metadata/met_corpus_principal_v2_revisado.csv",
    "secundario": "data/metadata/met_corpus_secundario_v2_revisado.csv",
    "descartados": "data/metadata/met_descartados_v2_revisado.csv",
}

SALIDA_PENDIENTES = "data/metadata/met_pendientes_revision_v2.csv"
REPORTE = "outputs/reports/met_resumen_revision_manual_v2.md"

CORPUS_VALIDOS = {"principal", "secundario", "descartados", "revisar"}

COLUMNAS_SALIDA = [
    "id_fuente",
    "corpus_actual",
    "corpus_final",
    "decision_revision",
    "motivo_revision",
    "observaciones_revision",
    "revisor",
    "fecha_revision",
    "titulo_original",
    "titulo_es_sugerido",
    "cultura",
    "fecha_objeto",
    "procedencia",
    "tipo_objeto",
    "tipo_superficie",
    "material",
    "material_normalizado",
    "tecnica",
    "estado_curacion",
    "destino_curacion",
    "motivo_curacion",
    "url_objeto",
    "url_imagen",
    "ruta_imagen_local",
    "licencia",
    "origen_curacion",
]


def texto(valor: object) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def leer_revision(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=True)

    if HOJA_REVISION not in wb.sheetnames:
        raise ValueError(f"No existe la hoja requerida: {HOJA_REVISION}")

    ws = wb[HOJA_REVISION]
    encabezados = [texto(cell.value) for cell in ws[1]]

    filas: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fila = {
            encabezados[idx]: texto(value)
            for idx, value in enumerate(row)
            if idx < len(encabezados)
        }

        if not fila.get("id_fuente"):
            continue

        filas.append(fila)

    return filas


def normalizar_corpus_final(fila: dict[str, str]) -> str:
    corpus_final = texto(fila.get("corpus_final")).lower()

    if not corpus_final:
        return texto(fila.get("corpus_actual")).lower()

    if corpus_final not in CORPUS_VALIDOS:
        return "revisar"

    return corpus_final


def preparar_fila_salida(fila: dict[str, str], corpus_final: str) -> dict[str, str]:
    salida = {columna: texto(fila.get(columna)) for columna in COLUMNAS_SALIDA}
    salida["corpus_final"] = corpus_final

    if corpus_final == "principal":
        salida["estado_curacion"] = "corpus_principal"
        salida["destino_curacion"] = "principal"
    elif corpus_final == "secundario":
        salida["estado_curacion"] = "corpus_secundario"
        salida["destino_curacion"] = "secundario"
    elif corpus_final == "descartados":
        salida["estado_curacion"] = "descartado"
        salida["destino_curacion"] = "exclusion_curatorial"

    return salida


def escribir_csv(path: Path, filas: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    with path.open("w", newline="", encoding="utf-8") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=COLUMNAS_SALIDA)
        writer.writeheader()
        writer.writerows(filas)


def escribir_reporte(
    path: Path,
    filas_revision: list[dict[str, str]],
    salidas: dict[str, list[dict[str, str]]],
    pendientes: list[dict[str, str]],
    problemas: list[str],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    movimientos = []
    for fila in filas_revision:
        corpus_actual = texto(fila.get("corpus_actual")).lower()
        corpus_final = normalizar_corpus_final(fila)

        if corpus_final != corpus_actual:
            movimientos.append((corpus_actual, corpus_final))

    conteo_movimientos = Counter(movimientos)

    lineas = [
        "# Resumen de revision manual MET v2",
        "",
        "Este reporte resume la aplicacion de decisiones manuales registradas en `data/metadata/met_revision_manual_v2.xlsx`.",
        "",
        "## Conteos finales",
        "",
        f"- Corpus principal revisado: {len(salidas['principal'])}",
        f"- Corpus secundario revisado: {len(salidas['secundario'])}",
        f"- Descartados revisados: {len(salidas['descartados'])}",
        f"- Pendientes por revisar: {len(pendientes)}",
        "",
        "## Movimientos registrados",
        "",
    ]

    if conteo_movimientos:
        for (origen, destino), cantidad in sorted(conteo_movimientos.items()):
            lineas.append(f"- {origen} -> {destino}: {cantidad}")
    else:
        lineas.append("- No se registraron movimientos entre corpus.")

    if pendientes:
        lineas.extend(["", "## Pendientes", ""])
        for fila in pendientes[:30]:
            lineas.append(
                f"- {fila.get('id_fuente', '')}: {fila.get('titulo_original', '')}"
            )

    if problemas:
        lineas.extend(["", "## Hallazgos", ""])
        for problema in problemas:
            lineas.append(f"- {problema}")

    path.write_text("\n".join(lineas) + "\n", encoding="utf-8")


def aplicar_revision(root: Path) -> int:
    workbook_path = root / WORKBOOK
    filas_revision = leer_revision(workbook_path)

    salidas = {
        "principal": [],
        "secundario": [],
        "descartados": [],
    }
    pendientes: list[dict[str, str]] = []
    problemas: list[str] = []

    ids_vistos: set[str] = set()
    ids_duplicados: set[str] = set()

    for fila in filas_revision:
        id_fuente = texto(fila.get("id_fuente"))

        if id_fuente in ids_vistos:
            ids_duplicados.add(id_fuente)
        ids_vistos.add(id_fuente)

        corpus_final = normalizar_corpus_final(fila)
        fila_salida = preparar_fila_salida(fila, corpus_final)

        if corpus_final == "revisar":
            pendientes.append(fila_salida)
        elif corpus_final in salidas:
            salidas[corpus_final].append(fila_salida)
        else:
            pendientes.append(fila_salida)
            problemas.append(f"{id_fuente}: corpus_final invalido: {corpus_final}")

    if ids_duplicados:
        muestra = ", ".join(sorted(ids_duplicados)[:20])
        problemas.append(f"IDs duplicados en workbook: {muestra}")

    for corpus, ruta_relativa in SALIDAS.items():
        escribir_csv(root / ruta_relativa, salidas[corpus])

    escribir_csv(root / SALIDA_PENDIENTES, pendientes)
    escribir_reporte(root / REPORTE, filas_revision, salidas, pendientes, problemas)

    print(f"Principal revisado: {len(salidas['principal'])}")
    print(f"Secundario revisado: {len(salidas['secundario'])}")
    print(f"Descartados revisados: {len(salidas['descartados'])}")
    print(f"Pendientes por revisar: {len(pendientes)}")
    print(f"Reporte: {root / REPORTE}")

    return 1 if problemas else 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Aplica la revision manual MET v2 desde Excel.")
    parser.add_argument("--root", default=".", help="Raiz del repositorio.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    return aplicar_revision(Path(args.root).resolve())


if __name__ == "__main__":
    raise SystemExit(main())