from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


COLUMNAS_REVISION = [
    "nro_revision",
    "fuente",
    "id_objeto",
    "numero_acceso",
    "titulo_original",
    "cultura",
    "periodo",
    "clasificacion_original",
    "nombre_objeto_original",
    "tecnica_original",
    "material_original",
    "descripcion_original",
    "url_objeto",
    "url_imagen",
    "enlace_objeto",
    "enlace_imagen",
    "terminos_andinos_auto",
    "terminos_textiles_auto",
    "categoria_revision_auto",
    "motivo_revision_auto",
    "decision_sugerida",
    "decision_auditoria",
    "motivo_auditoria",
    "observacion_auditoria",
    "incluir_en_corpus",
    "subconjunto_corpus",
    "calidad_visual",
    "tipo_objeto_manual",
    "cultura_manual",
    "periodo_manual",
    "notas_iconograficas",
    "notas_tecnicas",
    "revisor",
    "fecha_revision",
]

DECISIONES_AUDITORIA = [
    "aceptar_principal",
    "aceptar_secundario",
    "descartar_no_textil",
    "descartar_no_andino",
    "descartar_baja_calidad",
    "descartar_duplicado",
    "pendiente_discusion",
]

INCLUIR_EN_CORPUS = [
    "si",
    "no",
    "pendiente",
]

SUBCONJUNTO_CORPUS = [
    "principal",
    "secundario",
    "descartado",
    "pendiente",
]

CALIDAD_VISUAL = [
    "alta",
    "media",
    "baja",
    "no_aplica",
]

TIPOS_OBJETO = [
    "manto",
    "tunica",
    "camisa",
    "bolsa",
    "faja_o_banda",
    "fragmento",
    "textil_plumario",
    "otro_textil",
    "no_textil",
    "incierto",
]

COLOR_CABECERA = "1F4E78"
COLOR_MANUAL = "FFF2CC"
COLOR_AUTOMATICO = "E7E6E6"
COLOR_BORDE = "D9D9D9"
BLANCO = "FFFFFF"


def leer_filas_csv(ruta: Path, solo_candidatos: bool = True) -> list[dict[str, str]]:
    with ruta.open(newline="", encoding="utf-8-sig") as archivo:
        filas = list(csv.DictReader(archivo))

    if solo_candidatos:
        filas = [fila for fila in filas if fila.get("curation_status") == "candidate"]

    return filas


def texto_seguro(valor: object) -> str:
    if valor is None:
        return ""
    return " ".join(str(valor).split())


def estilo_cabecera(ws, fila: int) -> None:
    borde = Side(style="thin", color=COLOR_BORDE)

    for celda in ws[fila]:
        celda.font = Font(bold=True, color=BLANCO)
        celda.fill = PatternFill("solid", fgColor=COLOR_CABECERA)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = Border(top=borde, left=borde, right=borde, bottom=borde)


def estilo_cuerpo(ws, fila_inicio: int, fila_fin: int, col_fin: int) -> None:
    borde = Side(style="thin", color=COLOR_BORDE)

    for fila in ws.iter_rows(min_row=fila_inicio, max_row=fila_fin, max_col=col_fin):
        for celda in fila:
            celda.alignment = Alignment(vertical="top", wrap_text=True)
            celda.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    col_decision = COLUMNAS_REVISION.index("decision_auditoria") + 1

    for fila in ws.iter_rows(min_row=fila_inicio, max_row=fila_fin, min_col=1, max_col=col_decision - 1):
        for celda in fila:
            celda.fill = PatternFill("solid", fgColor=COLOR_AUTOMATICO)

    for fila in ws.iter_rows(min_row=fila_inicio, max_row=fila_fin, min_col=col_decision, max_col=col_fin):
        for celda in fila:
            celda.fill = PatternFill("solid", fgColor=COLOR_MANUAL)


def ajustar_anchos(ws) -> None:
    anchos = {
        "A": 10,
        "B": 24,
        "C": 14,
        "D": 18,
        "E": 36,
        "F": 22,
        "G": 18,
        "H": 22,
        "I": 24,
        "J": 26,
        "K": 34,
        "L": 48,
        "M": 30,
        "N": 30,
        "O": 18,
        "P": 18,
        "Q": 26,
        "R": 26,
        "S": 24,
        "T": 34,
        "U": 24,
        "V": 24,
        "W": 36,
        "X": 46,
        "Y": 18,
        "Z": 18,
        "AA": 16,
        "AB": 22,
        "AC": 22,
        "AD": 20,
        "AE": 38,
        "AF": 38,
        "AG": 18,
        "AH": 16,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho


def crear_hoja_vocabulario(wb: Workbook) -> None:
    ws = wb.create_sheet("vocabulario")

    vocabularios = {
        "decision_auditoria": DECISIONES_AUDITORIA,
        "incluir_en_corpus": INCLUIR_EN_CORPUS,
        "subconjunto_corpus": SUBCONJUNTO_CORPUS,
        "calidad_visual": CALIDAD_VISUAL,
        "tipo_objeto_manual": TIPOS_OBJETO,
    }

    for col_idx, (nombre, valores) in enumerate(vocabularios.items(), start=1):
        ws.cell(row=1, column=col_idx, value=nombre)
        ws.cell(row=1, column=col_idx).font = Font(bold=True)

        for fila_idx, valor in enumerate(valores, start=2):
            ws.cell(row=fila_idx, column=col_idx, value=valor)

        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 28

    ws.sheet_state = "hidden"


def agregar_validaciones(ws, fila_inicio: int, fila_fin: int) -> None:
    validaciones = {
        "decision_auditoria": "=vocabulario!$A$2:$A$8",
        "incluir_en_corpus": "=vocabulario!$B$2:$B$4",
        "subconjunto_corpus": "=vocabulario!$C$2:$C$5",
        "calidad_visual": "=vocabulario!$D$2:$D$5",
        "tipo_objeto_manual": "=vocabulario!$E$2:$E$11",
    }

    for nombre_columna, formula in validaciones.items():
        col_idx = COLUMNAS_REVISION.index(nombre_columna) + 1
        letra_col = ws.cell(row=1, column=col_idx).column_letter

        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Selecciona una opcion valida de la lista."
        dv.errorTitle = "Valor no permitido"
        dv.prompt = "Usa la lista desplegable."
        dv.promptTitle = nombre_columna
        ws.add_data_validation(dv)
        dv.add(f"{letra_col}{fila_inicio}:{letra_col}{fila_fin}")


def agregar_comentarios(ws) -> None:
    comentarios = {
        "decision_auditoria": "Decision manual final del revisor.",
        "motivo_auditoria": "Motivo breve que justifica la decision manual.",
        "observacion_auditoria": "Observaciones adicionales para trazabilidad.",
        "incluir_en_corpus": "Marcar si la pieza ingresa o no al corpus curado.",
        "subconjunto_corpus": "principal, secundario, descartado o pendiente.",
        "calidad_visual": "Evaluar utilidad de la imagen para tareas visuales.",
        "tipo_objeto_manual": "Tipo de objeto observado manualmente.",
        "cultura_manual": "Corregir o precisar cultura si corresponde.",
        "periodo_manual": "Corregir o precisar periodo si corresponde.",
        "notas_iconograficas": "Registrar motivos, figuras, patrones o composicion visual.",
        "notas_tecnicas": "Registrar tecnica, materialidad, tejido, bordado o tapiz.",
        "revisor": "Iniciales o nombre corto del revisor.",
        "fecha_revision": "Fecha en formato YYYY-MM-DD.",
    }

    for nombre_columna, comentario in comentarios.items():
        col_idx = COLUMNAS_REVISION.index(nombre_columna) + 1
        ws.cell(row=1, column=col_idx).comment = Comment(comentario, "UNI-CC")


def construir_hoja_revision(wb: Workbook, filas: list[dict[str, str]]) -> None:
    ws = wb.active
    ws.title = "revision_candidatos"

    ws.append(COLUMNAS_REVISION)

    for indice, fila in enumerate(filas, start=1):
        url_objeto = texto_seguro(fila.get("url"))
        url_imagen = texto_seguro(fila.get("image_url"))

        registro = {
            "nro_revision": indice,
            "fuente": texto_seguro(fila.get("source")),
            "id_objeto": texto_seguro(fila.get("source_id")),
            "numero_acceso": texto_seguro(fila.get("accession_number")),
            "titulo_original": texto_seguro(fila.get("title")),
            "cultura": texto_seguro(fila.get("culture")),
            "periodo": texto_seguro(fila.get("creation_date")),
            "clasificacion_original": texto_seguro(fila.get("classification")),
            "nombre_objeto_original": texto_seguro(fila.get("classification")),
            "tecnica_original": texto_seguro(fila.get("technique")),
            "material_original": texto_seguro(fila.get("medium")),
            "descripcion_original": texto_seguro(fila.get("description")),
            "url_objeto": url_objeto,
            "url_imagen": url_imagen,
            "enlace_objeto": "Abrir ficha",
            "enlace_imagen": "Abrir imagen",
            "terminos_andinos_auto": texto_seguro(fila.get("andean_terms")),
            "terminos_textiles_auto": texto_seguro(fila.get("textile_terms")),
            "categoria_revision_auto": texto_seguro(fila.get("curation_status")),
            "motivo_revision_auto": texto_seguro(fila.get("curation_notes")),
            "decision_sugerida": "revisar_candidato",
            "decision_auditoria": "",
            "motivo_auditoria": "",
            "observacion_auditoria": "",
            "incluir_en_corpus": "pendiente",
            "subconjunto_corpus": "pendiente",
            "calidad_visual": "",
            "tipo_objeto_manual": "",
            "cultura_manual": "",
            "periodo_manual": "",
            "notas_iconograficas": "",
            "notas_tecnicas": "",
            "revisor": "",
            "fecha_revision": "",
        }

        ws.append([registro[columna] for columna in COLUMNAS_REVISION])

        fila_excel = ws.max_row

        if url_objeto:
            ws.cell(row=fila_excel, column=COLUMNAS_REVISION.index("enlace_objeto") + 1).hyperlink = url_objeto
            ws.cell(row=fila_excel, column=COLUMNAS_REVISION.index("enlace_objeto") + 1).style = "Hyperlink"

        if url_imagen:
            ws.cell(row=fila_excel, column=COLUMNAS_REVISION.index("enlace_imagen") + 1).hyperlink = url_imagen
            ws.cell(row=fila_excel, column=COLUMNAS_REVISION.index("enlace_imagen") + 1).style = "Hyperlink"

    fila_fin = ws.max_row
    col_fin = len(COLUMNAS_REVISION)

    estilo_cabecera(ws, 1)
    estilo_cuerpo(ws, 2, fila_fin, col_fin)
    ajustar_anchos(ws)
    agregar_comentarios(ws)
    agregar_validaciones(ws, 2, fila_fin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=col_fin).column_letter}{fila_fin}"

    ref_tabla = f"A1:{ws.cell(row=fila_fin, column=col_fin).coordinate}"
    tabla = Table(displayName="RevisionCandidatosCMA", ref=ref_tabla)
    estilo_tabla = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tabla.tableStyleInfo = estilo_tabla
    ws.add_table(tabla)

    col_decision = ws.cell(row=1, column=COLUMNAS_REVISION.index("decision_auditoria") + 1).column_letter
    col_incluir = ws.cell(row=1, column=COLUMNAS_REVISION.index("incluir_en_corpus") + 1).column_letter

    ws.conditional_formatting.add(
        f"{col_decision}2:{col_decision}{fila_fin}",
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("descartar",{col_decision}2))'],
            fill=PatternFill("solid", fgColor="F4CCCC"),
        ),
    )

    ws.conditional_formatting.add(
        f"{col_incluir}2:{col_incluir}{fila_fin}",
        FormulaRule(
            formula=[f'{col_incluir}2="si"'],
            fill=PatternFill("solid", fgColor="D9EAD3"),
        ),
    )


def construir_hoja_instrucciones(wb: Workbook, total: int) -> None:
    ws = wb.create_sheet("instrucciones", 0)

    ws["A1"] = "Workbook de curacion manual CMA"
    ws["A1"].font = Font(bold=True, size=16, color=BLANCO)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_CABECERA)
    ws.merge_cells("A1:H1")

    ws["A2"] = (
        "Archivo generado automaticamente para revisar candidatos textiles andinos "
        "del Cleveland Museum of Art, manteniendo una estructura comparable con MET."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H2")

    ws.append([])
    ws.append(["Campo", "Descripcion"])

    instrucciones = [
        ("Insumo", "data/metadata/cma_andes_textiles_candidates.csv"),
        ("Registros incluidos", f"{total} candidatos con categoria_revision_auto = candidate"),
        ("Paso 1", "Abrir la ficha institucional y la imagen usando enlace_objeto y enlace_imagen."),
        ("Paso 2", "Completar decision_auditoria, motivo_auditoria y observacion_auditoria."),
        ("Paso 3", "Definir incluir_en_corpus y subconjunto_corpus."),
        ("Paso 4", "Completar calidad_visual y tipo_objeto_manual cuando corresponda."),
        ("Regla principal", "No ingresar al corpus principal piezas no textiles, no andinas, duplicadas o visualmente poco informativas."),
        ("Consistencia MET", "La estructura conserva los campos base usados en la auditoria MET: id_objeto, titulo_original, cultura, periodo, clasificacion_original, material_original, url_objeto, url_imagen, decision_auditoria, motivo_auditoria y observacion_auditoria."),
    ]

    for fila in instrucciones:
        ws.append(list(fila))

    estilo_cabecera(ws, 4)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110

    for fila in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=2):
        for celda in fila:
            celda.alignment = Alignment(wrap_text=True, vertical="top")


def construir_hoja_resumen(wb: Workbook, total: int) -> None:
    ws = wb.create_sheet("resumen", 1)

    ws["A1"] = "Resumen de curacion CMA"
    ws["A1"].font = Font(bold=True, size=16, color=BLANCO)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_CABECERA)
    ws.merge_cells("A1:D1")

    ws["A3"] = "Indicador"
    ws["B3"] = "Valor"
    estilo_cabecera(ws, 3)

    filas_resumen = [
        ("Total candidatos", total),
        ("Decisiones completadas", '=COUNTA(revision_candidatos!V:V)-1'),
        ("Pendientes decision_auditoria", '=COUNTBLANK(revision_candidatos!V:V)-1'),
        ("Incluir en corpus = si", '=COUNTIF(revision_candidatos!Y:Y,"si")'),
        ("Incluir en corpus = no", '=COUNTIF(revision_candidatos!Y:Y,"no")'),
        ("Incluir en corpus = pendiente", '=COUNTIF(revision_candidatos!Y:Y,"pendiente")'),
    ]

    for indicador, valor in filas_resumen:
        ws.append([indicador, valor])

    inicio = ws.max_row + 2
    ws.cell(row=inicio, column=1, value="decision_auditoria")
    ws.cell(row=inicio, column=2, value="conteo")
    estilo_cabecera(ws, inicio)

    for decision in DECISIONES_AUDITORIA:
        fila_destino = ws.max_row + 1
        ws.append([decision, f'=COUNTIF(revision_candidatos!V:V,A{fila_destino})'])

    inicio = ws.max_row + 2
    ws.cell(row=inicio, column=1, value="subconjunto_corpus")
    ws.cell(row=inicio, column=2, value="conteo")
    estilo_cabecera(ws, inicio)

    for subconjunto in SUBCONJUNTO_CORPUS:
        fila_destino = ws.max_row + 1
        ws.append([subconjunto, f'=COUNTIF(revision_candidatos!Z:Z,A{fila_destino})'])

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18


def escribir_reporte_markdown(ruta: Path, csv_entrada: Path, xlsx_salida: Path, total: int) -> None:
    ruta.parent.mkdir(parents=True, exist_ok=True)
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    texto = f"""# Workbook de curacion manual CMA

- Fecha de generacion: {fecha}
- CSV de entrada: `{csv_entrada}`
- Workbook generado: `{xlsx_salida}`
- Registros incluidos: {total}
- Criterio de inclusion: `curation_status = candidate`

## Hojas generadas

- `instrucciones`
- `resumen`
- `revision_candidatos`
- `vocabulario`

## Estandar aplicado

El workbook usa columnas en español y mantiene compatibilidad metodologica con la revision MET.
Las columnas manuales principales son:

- `decision_auditoria`
- `motivo_auditoria`
- `observacion_auditoria`
- `incluir_en_corpus`
- `subconjunto_corpus`
- `calidad_visual`
- `tipo_objeto_manual`
- `notas_iconograficas`
- `notas_tecnicas`
"""
    ruta.write_text(texto, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Crea workbook Excel en espanol para curacion manual de candidatos CMA."
    )
    parser.add_argument("--root", default=".", help="Raiz del repositorio.")
    parser.add_argument(
        "--csv",
        default="data/metadata/cma_andes_textiles_candidates.csv",
        help="CSV normalizado CMA.",
    )
    parser.add_argument(
        "--output",
        default="data/metadata/cma_revision_manual_textiles_andinos.xlsx",
        help="Workbook Excel de salida.",
    )
    parser.add_argument(
        "--include-all",
        action="store_true",
        help="Incluye tambien registros no candidatos.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = Path(args.root).resolve()

    csv_entrada = root / args.csv
    xlsx_salida = root / args.output
    reporte_salida = root / "outputs/reports/cma_revision_manual_summary.md"

    if not csv_entrada.exists():
        raise FileNotFoundError(f"No existe el CSV de entrada: {csv_entrada}")

    filas = leer_filas_csv(csv_entrada, solo_candidatos=not args.include_all)

    xlsx_salida.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    construir_hoja_revision(wb, filas)
    crear_hoja_vocabulario(wb)
    construir_hoja_resumen(wb, len(filas))
    construir_hoja_instrucciones(wb, len(filas))

    wb.save(xlsx_salida)
    escribir_reporte_markdown(reporte_salida, csv_entrada, xlsx_salida, len(filas))

    print(f"Workbook generado: {xlsx_salida}")
    print(f"Registros incluidos: {len(filas)}")
    print(f"Reporte: {reporte_salida}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())