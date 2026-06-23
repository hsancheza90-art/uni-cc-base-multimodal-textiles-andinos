from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


INPUT_SAMPLE = Path("data/metadata/corpus_piloto_annotation_sample.csv")
OUTPUT_XLSX = Path("data/metadata/corpus_piloto_annotation_sheet.xlsx")


ANNOTATION_COLUMNS = [
    "nro",
    "id_objeto",
    "titulo_original",
    "cultura",
    "periodo",
    "tipo_objeto",
    "tipo_superficie",
    "imagen_url",
    "url",
    "color_dominante",
    "colores",
    "contraste",
    "densidad_visual",
    "motivos",
    "familia_iconografica",
    "simetria",
    "composicion",
    "estado_anotacion",
    "observaciones_anotacion",
]


VOCABULARY = {
    "color_dominante": [
        "rojo",
        "azul",
        "amarillo",
        "negro",
        "blanco",
        "marron",
        "beige",
        "verde",
        "naranja",
        "multicolor",
        "no_determinado",
    ],
    "contraste": [
        "alto",
        "medio",
        "bajo",
        "no_determinado",
    ],
    "densidad_visual": [
        "alta",
        "media",
        "baja",
        "no_determinado",
    ],
    "familia_iconografica": [
        "geometrica",
        "zoomorfa",
        "antropomorfa",
        "fitomorfa",
        "abstracta",
        "mixta",
        "no_determinada",
    ],
    "simetria": [
        "bilateral",
        "radial",
        "repetitiva",
        "sin_simetria_clara",
        "no_determinado",
    ],
    "composicion": [
        "central",
        "modular",
        "franjas",
        "cuadricula",
        "paneles",
        "repeticion_lineal",
        "composicion_libre",
        "no_determinado",
    ],
    "estado_anotacion": [
        "pendiente",
        "anotado",
        "revisar",
        "descartar",
    ],
}


def build_annotation_workbook() -> None:
    if not INPUT_SAMPLE.exists():
        raise FileNotFoundError(f"No existe el archivo: {INPUT_SAMPLE}")

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(INPUT_SAMPLE)

    if "nro" not in df.columns:
        df.insert(0, "nro", range(1, len(df) + 1))

    annotation_cols = [col for col in ANNOTATION_COLUMNS if col in df.columns]
    annotation_df = df[annotation_cols].copy()

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        annotation_df.to_excel(writer, sheet_name="Anotacion", index=False)
        df.to_excel(writer, sheet_name="Referencia", index=False)

        vocab_rows = []
        max_len = max(len(values) for values in VOCABULARY.values())

        for i in range(max_len):
            row = {}
            for field, values in VOCABULARY.items():
                row[field] = values[i] if i < len(values) else ""
            vocab_rows.append(row)

        vocab_df = pd.DataFrame(vocab_rows)
        vocab_df.to_excel(writer, sheet_name="Vocabulario", index=False)

    format_workbook()

    print(f"Archivo Excel de anotación generado: {OUTPUT_XLSX}")


def format_workbook() -> None:
    wb = load_workbook(OUTPUT_XLSX)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        ws.auto_filter.ref = ws.dimensions

    ws = wb["Anotacion"]

    widths = {
        "A": 6,    # nro
        "B": 12,   # id_objeto
        "C": 32,   # titulo
        "D": 16,   # cultura
        "E": 14,   # periodo
        "F": 22,   # tipo_objeto
        "G": 28,   # tipo_superficie
        "H": 18,   # imagen_url
        "I": 16,   # url
        "J": 18,   # color_dominante
        "K": 26,   # colores
        "L": 16,   # contraste
        "M": 18,   # densidad
        "N": 30,   # motivos
        "O": 22,   # familia
        "P": 20,   # simetria
        "Q": 22,   # composicion
        "R": 18,   # estado
        "S": 38,   # observaciones
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 48

    # Convertir URLs en hyperlinks más fáciles de abrir
    headers = {cell.value: cell.column for cell in ws[1]}

    for col_name in ["imagen_url", "url"]:
        if col_name in headers:
            col_idx = headers[col_name]

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)

                if cell.value:
                    cell.hyperlink = cell.value
                    cell.value = "abrir"
                    cell.style = "Hyperlink"

    # Validaciones tipo lista
    vocab_ws = wb["Vocabulario"]
    vocab_headers = {cell.value: cell.column_letter for cell in vocab_ws[1]}

    for field, values in VOCABULARY.items():
        if field not in headers or field not in vocab_headers:
            continue

        target_col = headers[field]
        vocab_col = vocab_headers[field]
        formula = f"=Vocabulario!${vocab_col}$2:${vocab_col}${len(values) + 1}"

        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Selecciona un valor del vocabulario controlado."
        dv.errorTitle = "Valor no permitido"
        dv.prompt = "Selecciona un valor de la lista."
        dv.promptTitle = field

        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=target_col).coordinate}:{ws.cell(row=ws.max_row, column=target_col).coordinate}")

    # Colorear columnas editables
    editable_fill = PatternFill("solid", fgColor="FFF2CC")

    editable_cols = [
        "color_dominante",
        "colores",
        "contraste",
        "densidad_visual",
        "motivos",
        "familia_iconografica",
        "simetria",
        "composicion",
        "estado_anotacion",
        "observaciones_anotacion",
    ]

    for field in editable_cols:
        if field in headers:
            col_idx = headers[field]
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).fill = editable_fill

    # Ocultar hoja de referencia si prefieres una interfaz limpia
    # wb["Referencia"].sheet_state = "hidden"

    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    build_annotation_workbook()