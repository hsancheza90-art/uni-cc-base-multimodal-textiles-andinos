from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ENTRADAS = [
    ("principal", "data/metadata/met_corpus_principal_v2.csv"),
    ("secundario", "data/metadata/met_corpus_secundario_v2.csv"),
    ("descartados", "data/metadata/met_descartados_v2.csv"),
]

COLUMNAS = [
    "id_fuente",
    "corpus_actual",
    "corpus_final",
    "decision_revision",
    "motivo_revision",
    "observaciones_revision",
    "revisor",
    "fecha_revision",
    "titulo_original",
    "titulo_es_sugerido",
    "cultura",
    "fecha_objeto",
    "procedencia",
    "tipo_objeto",
    "tipo_superficie",
    "material",
    "material_normalizado",
    "tecnica",
    "estado_curacion",
    "destino_curacion",
    "motivo_curacion",
    "url_objeto",
    "url_imagen",
    "ruta_imagen_local",
    "licencia",
    "origen_curacion",
]

OPCIONES_CORPUS_FINAL = [
    "principal",
    "secundario",
    "descartados",
    "revisar",
]

OPCIONES_DECISION = [
    "mantener",
    "mover",
    "descartar",
    "recuperar",
    "revisar",
]


def leer_csv(ruta: Path) -> list[dict[str, str]]:
    with ruta.open(newline="", encoding="utf-8-sig") as archivo:
        return list(csv.DictReader(archivo))


def valor(fila: dict[str, str], campo: str) -> str:
    return (fila.get(campo) or "").strip()


def construir_filas(raiz: Path) -> list[dict[str, str]]:
    filas_revision: list[dict[str, str]] = []

    for corpus_actual, ruta_relativa in ENTRADAS:
        filas = leer_csv(raiz / ruta_relativa)

        for fila in filas:
            salida = {
                "id_fuente": valor(fila, "id_fuente"),
                "corpus_actual": corpus_actual,
                "corpus_final": corpus_actual,
                "decision_revision": "mantener",
                "motivo_revision": "",
                "observaciones_revision": "",
                "revisor": "",
                "fecha_revision": "",
                "titulo_original": valor(fila, "titulo_original"),
                "titulo_es_sugerido": valor(fila, "titulo_es_sugerido"),
                "cultura": valor(fila, "cultura"),
                "fecha_objeto": valor(fila, "fecha_objeto"),
                "procedencia": valor(fila, "procedencia"),
                "tipo_objeto": valor(fila, "tipo_objeto"),
                "tipo_superficie": valor(fila, "tipo_superficie"),
                "material": valor(fila, "material"),
                "material_normalizado": valor(fila, "material_normalizado"),
                "tecnica": valor(fila, "tecnica"),
                "estado_curacion": valor(fila, "estado_curacion"),
                "destino_curacion": valor(fila, "destino_curacion"),
                "motivo_curacion": valor(fila, "motivo_curacion"),
                "url_objeto": valor(fila, "url_objeto"),
                "url_imagen": valor(fila, "url_imagen"),
                "ruta_imagen_local": valor(fila, "ruta_imagen_local"),
                "licencia": valor(fila, "licencia"),
                "origen_curacion": valor(fila, "origen_curacion"),
            }
            filas_revision.append(salida)

    return filas_revision


def aplicar_estilo_revision(ws) -> None:
    fill_header = PatternFill("solid", fgColor="1F4E78")
    font_header = Font(color="FFFFFF", bold=True)

    for celda in ws[1]:
        celda.fill = fill_header
        celda.font = font_header
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for fila in ws.iter_rows(min_row=2):
        for celda in fila:
            celda.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    anchos = {
        "A": 14,
        "B": 16,
        "C": 16,
        "D": 18,
        "E": 42,
        "F": 46,
        "G": 18,
        "H": 18,
        "I": 34,
        "J": 34,
        "K": 20,
        "L": 20,
        "M": 24,
        "N": 20,
        "O": 24,
        "P": 34,
        "Q": 28,
        "R": 18,
        "S": 18,
        "T": 20,
        "U": 58,
        "V": 42,
        "W": 42,
        "X": 36,
        "Y": 34,
        "Z": 24,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho


def agregar_validaciones(ws) -> None:
    max_row = max(ws.max_row, 2)

    dv_corpus = DataValidation(
        type="list",
        formula1='"' + ",".join(OPCIONES_CORPUS_FINAL) + '"',
        allow_blank=False,
    )
    dv_decision = DataValidation(
        type="list",
        formula1='"' + ",".join(OPCIONES_DECISION) + '"',
        allow_blank=False,
    )

    ws.add_data_validation(dv_corpus)
    ws.add_data_validation(dv_decision)

    dv_corpus.add(f"C2:C{max_row}")
    dv_decision.add(f"D2:D{max_row}")


def agregar_hipervinculos(ws) -> None:
    # V = url_objeto, W = url_imagen segun COLUMNAS.
    for row in range(2, ws.max_row + 1):
        url_objeto = ws[f"V{row}"].value
        url_imagen = ws[f"W{row}"].value

        if url_objeto:
            ws[f"V{row}"].hyperlink = url_objeto
            ws[f"V{row}"].style = "Hyperlink"

        if url_imagen:
            ws[f"W{row}"].hyperlink = url_imagen
            ws[f"W{row}"].style = "Hyperlink"


def crear_hoja_revision(wb: Workbook, filas: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("revision_met_v2")
    ws.append(COLUMNAS)

    for fila in filas:
        ws.append([fila.get(columna, "") for columna in COLUMNAS])

    aplicar_estilo_revision(ws)
    agregar_validaciones(ws)
    agregar_hipervinculos(ws)


def crear_hoja_listas(wb: Workbook) -> None:
    ws = wb.create_sheet("listas")
    ws["A1"] = "corpus_final"
    ws["B1"] = "decision_revision"

    for idx, opcion in enumerate(OPCIONES_CORPUS_FINAL, start=2):
        ws[f"A{idx}"] = opcion

    for idx, opcion in enumerate(OPCIONES_DECISION, start=2):
        ws[f"B{idx}"] = opcion

    for celda in ws[1]:
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor="D9EAF7")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24


def crear_hoja_resumen(wb: Workbook, filas: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("resumen", 0)

    conteos: dict[str, int] = {}
    for fila in filas:
        corpus = fila["corpus_actual"]
        conteos[corpus] = conteos.get(corpus, 0) + 1

    ws["A1"] = "Revision manual MET v2"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Corpus actual"
    ws["B3"] = "Registros"
    ws["A3"].font = Font(bold=True)
    ws["B3"].font = Font(bold=True)

    fila_excel = 4
    for corpus in ["principal", "secundario", "descartados"]:
        ws[f"A{fila_excel}"] = corpus
        ws[f"B{fila_excel}"] = conteos.get(corpus, 0)
        fila_excel += 1

    ws["A9"] = "Uso"
    ws["A9"].font = Font(bold=True)
    ws["A10"] = "Editar corpus_final para mover un registro entre principal, secundario, descartados o revisar."
    ws["A11"] = "corpus_actual conserva la clasificacion base auditada y no debe modificarse."
    ws["A12"] = "decision_revision y motivo_revision documentan la decision humana."

    ws.column_dimensions["A"].width = 96
    ws.column_dimensions["B"].width = 18


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Crea el workbook maestro de revision manual MET v2.")
    parser.add_argument("--root", default=".", help="Raiz del repositorio.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    raiz = Path(args.root).resolve()

    filas = construir_filas(raiz)

    wb = Workbook()
    wb.remove(wb.active)

    crear_hoja_resumen(wb, filas)
    crear_hoja_revision(wb, filas)
    crear_hoja_listas(wb)

    salida = raiz / "data/metadata/met_revision_manual_v2.xlsx"
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)

    print(f"Workbook creado: {salida}")
    print(f"Registros para revision: {len(filas)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())